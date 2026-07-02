"""
ExcelファイルをZEATSのJSONデータに変換するモジュール。

対応フォーマット: 大阪フォーマット.xlsx
  - シートごとに座席ブロックが横方向に展開
  - セルの値=座席番号(整数)、背景色=販売チャネル
  - C列に行ラベル('N列')、'ブロック'含む文字列セルがブロックヘッダー

色マッチング戦略:
  1. 凡例A列の色と座席色が完全一致 → そのまま使用
  2. 不一致の場合（アウトレットサイドの事務局など）→ 凡例の枚数と
     座席色の出現数が一致するカウントベースマッチングで補完
  3. 凡例に存在しない独自色（テーブル席など）→ EXTRA_COLOR_MAPPINGS で補完
"""

import json
import re
from collections import Counter
import openpyxl

SKIP_SHEETS: set[str] = set()  # 座席0件のシートは parse_excel 内で自動スキップ

# 凡例に記載のない独自色とチャネルIDの対応（シートをまたいで適用）
# テーブル席では標準凡例と異なる色で事務局・ふるさと納税が使われている
EXTRA_COLOR_MAPPINGS: dict[str, str] = {
    'FFFAE2D5': '事務局',       # テーブル席の事務局色
    'FFFBE2D5': 'ふるさと納税',  # テーブル席のふるさと納税色
}

SHEET_SHORT_NAMES = {
    'マーブルビーチサイド': 'マーブル',
    'アウトレットサイド':   'アウトレット',
    'テーブル席':           'テーブル',
    'テーブル席（追加）':   'テーブル追加',
}


def _is_seat_value(v) -> bool:
    """セル値が座席番号かどうか判定。int/floatを許容し bool を除外する。"""
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def _rgb(cell) -> str | None:
    """セルの背景色をARGB文字列で返す。塗りなし・透明はNone。"""
    if cell.fill and cell.fill.patternType not in (None, 'none'):
        rgb = cell.fill.fgColor.rgb
        if rgb and rgb not in ('00000000', ''):
            return rgb
    return None


def _parse_count(value) -> int | None:
    """凡例B列の値を整数に変換。数式文字列は評価せずNoneを返す。"""
    if isinstance(value, (int, float)):
        return int(value)
    return None  # 数式文字列('=...')はカウントなし扱い


def _find_legend(ws) -> tuple[list[dict], dict[str, str]]:
    """
    凡例行を走査しチャネル情報を返す。
    色付きモード: A列に色 + B列に数値/数式
    白図面モード: A列にテキスト + C列に「N列」パターン（色なし・B列なし）
    戻り値: (categories_list, {legend_rgb: channel_id})
    """
    # 白図面モード用: C列に「N列」パターンがある行番号を収集
    row_label_rows: set[int] = set()
    for cell in ws['C']:
        if cell.value and isinstance(cell.value, str):
            if re.fullmatch(r'\d+列', cell.value.strip()):
                row_label_rows.add(cell.row)

    categories = []
    color_to_id: dict[str, str] = {}
    seen_ids: set[str] = set()

    for row in ws.iter_rows(max_col=3):
        a = row[0]
        b = row[1] if len(row) > 1 else None
        if not (a.value and isinstance(a.value, str)):
            continue

        rgb = _rgb(a)
        b_val = b.value if b else None
        b_is_count = (isinstance(b_val, (int, float)) or
                      (isinstance(b_val, str) and b_val.startswith('=')))

        channel_id = a.value.strip()
        if channel_id in seen_ids:
            continue

        if rgb and b_is_count:
            # 色付きモード: 従来通り色+枚数で凡例検出
            count = _parse_count(b_val)
            hex_color = '#' + rgb[2:]
            categories.append({
                'id':     channel_id,
                'name':   channel_id,
                'color':  hex_color,
                '_rgb':   rgb,
                '_count': count,
            })
            color_to_id[rgb] = channel_id
            seen_ids.add(channel_id)
        elif not rgb and not b_is_count and a.row in row_label_rows:
            # 白図面モード: 色なし・B列なし・C列に行ラベルあり → チャネル名行
            categories.append({
                'id':     channel_id,
                'name':   channel_id,
                'color':  None,
                '_rgb':   None,
                '_count': None,
            })
            seen_ids.add(channel_id)
        elif rgb and not b_is_count:
            # 色のみ凡例モード: A列に色あり・B列に枚数なし
            # 座席セルには色がないため座席割当はされず、区分色のみ取得する
            hex_color = '#' + rgb[2:]
            categories.append({
                'id':     channel_id,
                'name':   channel_id,
                'color':  hex_color,
                '_rgb':   rgb,
                '_count': None,
            })
            color_to_id[rgb] = channel_id
            seen_ids.add(channel_id)

    return categories, color_to_id


def _build_color_map(ws, categories: list[dict],
                     legend_color_to_id: dict[str, str]) -> dict[str, str]:
    """
    座席セルの色と凡例を突き合わせて完全なcolor→idマップを返す。

    1次: 凡例A列の色と座席色が完全一致
    2次（フォールバック）: 凡例A列の色が座席に存在しない場合、
                           凡例の枚数と座席色の出現数が完全一致するものを照合
                           ※アウトレットサイドの事務局色ズレへの対処
    """
    # 全座席セルの色をカウント
    seat_color_counts: Counter = Counter()
    for row in ws.iter_rows():
        for cell in row:
            if _is_seat_value(cell.value):
                rgb = _rgb(cell)
                if rgb:
                    seat_color_counts[rgb] += 1

    color_to_id = dict(legend_color_to_id)

    # 凡例色が実際の座席に存在するかチェック
    legend_colors_in_seats = set(seat_color_counts.keys()) & set(legend_color_to_id.keys())

    # 未マッチの座席色（凡例にない色）
    unmatched_seat_colors = {c: n for c, n in seat_color_counts.items()
                             if c not in color_to_id}

    # フォールバック: 凡例色が座席に存在しないチャネルをカウントで補完
    for cat in categories:
        if cat['_rgb'] in legend_colors_in_seats:
            continue  # 色で既にマッチ済み
        count = cat['_count']
        if count is None:
            continue
        exact = [c for c, n in unmatched_seat_colors.items() if n == count]
        if len(exact) == 1:
            color_to_id[exact[0]] = cat['id']
            del unmatched_seat_colors[exact[0]]

    # 3次: EXTRA_COLOR_MAPPINGS で残った未マッチ色を補完
    for rgb, channel_id in EXTRA_COLOR_MAPPINGS.items():
        if rgb not in color_to_id:
            color_to_id[rgb] = channel_id

    return color_to_id


def _find_block_headers(ws) -> list[tuple[int, int, str]]:
    """'ブロック'を含む文字列セルを走査して (row, col, name) のリストを返す。"""
    headers = []
    for row in ws.iter_rows():
        for cell in row:
            if (cell.value and isinstance(cell.value, str)
                    and 'ブロック' in cell.value):
                headers.append((cell.row, cell.column, cell.value.strip()))
    return headers


def _find_row_labels(ws) -> dict[int, int]:
    """C列の 'N列' パターンセルから {excel_row: 列番号} を返す。"""
    labels: dict[int, int] = {}
    for cell in ws['C']:
        if cell.value and isinstance(cell.value, str):
            m = re.fullmatch(r'(\d+)列', cell.value.strip())
            if m:
                labels[cell.row] = int(m.group(1))
    return labels


def _find_section_headers(ws) -> list[tuple[int, str]]:
    """
    A列のテキストセルから (row, section_name) を返す。
    凡例行（B列に数値/式あり）は除外。
    """
    headers = []
    for row in ws.iter_rows(max_col=2):
        a = row[0]
        b = row[1] if len(row) > 1 else None
        if not (a.value and isinstance(a.value, str)):
            continue
        b_val = b.value if b else None
        b_is_count = (isinstance(b_val, (int, float)) or
                      (isinstance(b_val, str) and b_val.startswith('=')))
        if b_is_count:
            continue
        headers.append((a.row, a.value.strip()))
    return headers


def _resolve_block(cell_row: int, cell_col: int,
                   block_headers: list[tuple[int, int, str]],
                   section_headers: list[tuple[int, str]]) -> str | None:
    """座席セルに対して最も近い上・左のブロック名を返す。"""
    # ブロックヘッダー優先: col ≤ cell_col AND row ≤ cell_row で最近傍
    candidates = [(r, c, n) for r, c, n in block_headers
                  if c <= cell_col and r <= cell_row]
    if candidates:
        max_r = max(x[0] for x in candidates)
        same_r = [x for x in candidates if x[0] == max_r]
        return max(same_r, key=lambda x: x[1])[2]

    # セクションヘッダーで代替: row ≤ cell_row で最近傍
    candidates_s = [(r, n) for r, n in section_headers if r <= cell_row]
    if candidates_s:
        return max(candidates_s, key=lambda x: x[0])[1]

    return None


def _parse_sheet(ws, short_name: str,
                 color_to_id: dict[str, str]) -> dict[str, dict]:
    """1シートを走査して seats 辞書を返す。{seat_id: {categoryId, ticketNo}}"""
    block_headers   = _find_block_headers(ws)
    row_labels      = _find_row_labels(ws)
    section_headers = _find_section_headers(ws)
    seats: dict[str, dict] = {}

    for row in ws.iter_rows():
        for cell in row:
            if cell.column <= 3:
                continue  # A列=セクション名, B列=凡例枚数, C列=行ラベル → 座席対象外
            if not _is_seat_value(cell.value):
                continue
            row_num = row_labels.get(cell.row)
            if row_num is None:
                continue
            block = _resolve_block(cell.row, cell.column,
                                   block_headers, section_headers)
            if block is None:
                continue

            seat_num = int(cell.value)
            seat_id = f"{short_name}_{block}_{row_num}列_{seat_num}番"
            rgb = _rgb(cell)
            channel_id = color_to_id.get(rgb) if rgb else None

            seats[seat_id] = {
                'categoryId': channel_id,
                'ticketNo':   None,
                '_col':       cell.column,  # Excelの列位置（グリッド配置用・出力時に除去）
                '_row':       cell.row,     # Excelの行位置（ティア分離用・出力時に除去）
            }

    return seats


def _generate_blocks(seats: dict, short_name: str,
                     block_headers: list[tuple[int, int, str]]) -> list[dict]:
    """
    seats辞書からブロック別の表示構造を生成する。

    Excelの実際の行位置（_row）をキーとして管理し、同一ブロック内に複数の
    ティア（行番号リセット）がある場合も正しく再現する。

    block_headers: _find_block_headers() の戻り値。ブロックのヘッダー行・列位置を
                   _tier / _hpos に使用することで、アプリ上でのブロック横並び順を
                   Excel と一致させる。

    戻り値:
      [ { "name":    block_name,
          "rowNums": [row_num, ...],  # 行の「N列」番号（C列ラベル）
          "rows":    [ [seat_id|None, ...] ],
          "_tier":   int,             # 横並びグループ（ブロックヘッダーの行番号）
          "_hpos":   int,             # ティア内の横順（ブロックヘッダーの列番号）
        }, ... ]
    """
    prefix = short_name + '_'
    # ブロック名 → (header_row, header_col)
    header_pos: dict[str, tuple[int, int]] = {
        name: (row, col) for row, col, name in block_headers
    }

    # block_seat_rows[block][excel_row][col] = seat_id
    block_seat_rows: dict[str, dict[int, dict[int, str]]] = {}
    # block_row_labels[block][excel_row] = row_num (C列ラベル)
    block_row_labels: dict[str, dict[int, int]] = {}

    for seat_id, seat_info in seats.items():
        rest = seat_id[len(prefix):]
        m = re.search(r'^(.+)_(\d+)列_(\d+)番$', rest)
        if not m:
            continue
        block     = m.group(1)
        row_num   = int(m.group(2))
        col       = seat_info.get('_col', 0)
        excel_row = seat_info.get('_row', 0)
        block_seat_rows.setdefault(block, {}).setdefault(excel_row, {})[col] = seat_id
        block_row_labels.setdefault(block, {})[excel_row] = row_num

    def _natural_key(s: str):
        return [int(t) if t.isdigit() else t for t in re.split(r'(\d+)', s)]

    blocks = []
    for block_name in sorted(block_seat_rows, key=_natural_key):
        rows_dict = block_seat_rows[block_name]  # {excel_row: {col: seat_id}}
        labels    = block_row_labels[block_name]  # {excel_row: row_num}

        all_cols = [c for cols in rows_dict.values() for c in cols]
        if not all_cols:
            continue
        min_col, max_col = min(all_cols), max(all_cols)

        sorted_excel_rows = sorted(rows_dict)
        row_nums = [labels[er] for er in sorted_excel_rows]
        rows = [
            [rows_dict[er].get(c) for c in range(min_col, max_col + 1)]
            for er in sorted_excel_rows
        ]

        # ヘッダー行・列をティア情報として使用（フォールバック: 座席の最小行・列）
        hdr_row, hdr_col = header_pos.get(block_name,
                                          (min(sorted_excel_rows), min_col))
        blocks.append({
            'name':    block_name,
            'rowNums': row_nums,
            'rows':    rows,
            '_tier':   hdr_row,   # ブロックヘッダーのExcel行 → 横並びグループ
            '_hpos':   hdr_col,   # ブロックヘッダーのExcel列 → ティア内の横順
        })

    return blocks


def parse_excel(filepath: str) -> dict:
    """
    ExcelファイルをZEATSのJSONデータに変換する。

    戻り値:
    {
      "version": 1,
      "categories": [{id, name, color}, ...],
      "sheets": {
        "シート名": {
          "name": str,
          "shortName": str,
          "seats": {seat_id: {categoryId, ticketNo}}
        }
      }
    }
    """
    # data_only=Falseで読み込み（数式文字列として取得し凡例検出を安定化）
    wb = openpyxl.load_workbook(filepath, data_only=False)

    # " (数字)" サフィックスのシートは、同名のベースシートが存在する場合はスキップ
    # （例: 「マーブルビーチサイド (2)」があり「マーブルビーチサイド」も存在する旧バージョンシート）
    _duplicate_re = re.compile(r'\s+\(\d+\)$')
    _base_names = set(wb.sheetnames)
    def _is_duplicate_sheet(name: str) -> bool:
        m = _duplicate_re.search(name)
        return bool(m) and name[:m.start()] in _base_names

    # 凡例を持つ最初のシートからグローバルチャネル情報を取得
    global_categories: list[dict] = []
    global_legend_color_to_id: dict[str, str] = {}
    for _n in wb.sheetnames:
        if _n in SKIP_SHEETS or _is_duplicate_sheet(_n):
            continue
        _cats, _color_map = _find_legend(wb[_n])
        if _cats:
            global_categories = _cats
            global_legend_color_to_id = _color_map
            break

    # _countなどの内部フィールドを除いた公開用categories（color=Noneは省略）
    public_categories = [
        {k: v for k, v in {'id': c['id'], 'name': c['name'], 'color': c['color']}.items()
         if v is not None}
        for c in global_categories
    ]

    result: dict = {
        'version':    1,
        'categories': public_categories,
        'sheets':     {},
    }

    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS or _is_duplicate_sheet(sheet_name):
            continue

        ws = wb[sheet_name]
        short_name = SHEET_SHORT_NAMES.get(sheet_name, sheet_name)

        # シート固有の凡例を試みる
        sheet_cats, legend_color_to_id = _find_legend(ws)
        if sheet_cats:
            # 凡例あり: シート固有のフォールバックマッチを適用
            color_to_id = _build_color_map(ws, sheet_cats, legend_color_to_id)
        else:
            # 凡例なし（テーブル席など）: グローバルマップ + EXTRA_COLOR_MAPPINGS を使用
            color_to_id = {**global_legend_color_to_id, **EXTRA_COLOR_MAPPINGS}

        seats = _parse_sheet(ws, short_name, color_to_id)
        if not seats:
            continue  # 座席0件のシート（全体図・旧バージョンシート等）はスキップ
        block_headers = _find_block_headers(ws)
        blocks = _generate_blocks(seats, short_name, block_headers)

        # _col など内部フィールドを除去して公開用 seats を作成
        public_seats = {
            sid: {'categoryId': s['categoryId'], 'ticketNo': s['ticketNo']}
            for sid, s in seats.items()
        }

        result['sheets'][sheet_name] = {
            'name':      sheet_name,
            'shortName': short_name,
            'seats':     public_seats,
            'blocks':    blocks,
        }

    return result


def get_seat_cell_map(filepath: str) -> dict[str, tuple[str, int, int]]:
    """
    Excelファイルを再走査し {seat_id: (sheet_name, row, col)} を返す。
    seat_colorizer から座席セルの位置を特定するために使用する。
    """
    wb = openpyxl.load_workbook(filepath, data_only=False)
    _duplicate_re = re.compile(r'\s+\(\d+\)$')
    _base_names = set(wb.sheetnames)

    def _is_dup(name: str) -> bool:
        m = _duplicate_re.search(name)
        return bool(m) and name[:m.start()] in _base_names

    cell_map: dict[str, tuple[str, int, int]] = {}
    for sheet_name in wb.sheetnames:
        if _is_dup(sheet_name):
            continue
        ws = wb[sheet_name]
        short_name     = SHEET_SHORT_NAMES.get(sheet_name, sheet_name)
        block_headers  = _find_block_headers(ws)
        row_labels     = _find_row_labels(ws)
        section_headers = _find_section_headers(ws)

        for row in ws.iter_rows():
            for cell in row:
                if cell.column <= 3 or not _is_seat_value(cell.value):
                    continue
                row_num = row_labels.get(cell.row)
                if row_num is None:
                    continue
                block = _resolve_block(cell.row, cell.column,
                                       block_headers, section_headers)
                if block is None:
                    continue
                seat_id = f"{short_name}_{block}_{row_num}列_{int(cell.value)}番"
                cell_map[seat_id] = (sheet_name, cell.row, cell.column)

    return cell_map


if __name__ == '__main__':
    import sys

    filepath   = sys.argv[1] if len(sys.argv) > 1 else '../tmp/大阪フォーマット_白図面.xlsx'
    output     = sys.argv[2] if len(sys.argv) > 2 else None
    data = parse_excel(filepath)

    print(f"チャネル数: {len(data['categories'])}")
    for cat in data['categories']:
        print(f"  {cat['name']} ({cat.get('color', '色未設定')})")

    total_seats = 0
    for sheet_name, sheet in data['sheets'].items():
        seats = sheet['seats']
        total_seats += len(seats)
        channel_counts: dict[str, int] = {}
        for s in seats.values():
            key = s['categoryId'] or '未割当'
            channel_counts[key] = channel_counts.get(key, 0) + 1
        print(f"\n{sheet_name}: {len(seats)}席")
        for ch, cnt in sorted(channel_counts.items(), key=lambda x: -x[1]):
            print(f"  {ch}: {cnt}席")

    print(f"\n合計: {total_seats}席")

    out = output or 'output.json'
    with open(out, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"\nJSON出力: {out}")
