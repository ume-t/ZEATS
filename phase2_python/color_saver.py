"""
UIで設定した区分色を元Excelの凡例セル（A列の区分名セル）に適用して保存する。
"""
import shutil
import openpyxl
from openpyxl.styles import PatternFill


def _hex_to_argb(hex_color: str) -> str:
    """'#RRGGBB' → 'FFRRGGBB' (ARGB)"""
    h = hex_color.lstrip('#')
    if len(h) == 3:
        h = ''.join(c * 2 for c in h)
    return ('FF' + h).upper()


def apply_colors_to_excel(src_path: str, dst_path: str, color_map: dict) -> None:
    """
    src_path を dst_path にコピーし、A列の区分名セルに color_map の色を適用する。

    color_map: {channel_id: '#RRGGBB'}
    元ファイルは変更しない。
    """
    shutil.copy2(src_path, dst_path)
    wb = openpyxl.load_workbook(dst_path)

    for ws in wb.worksheets:
        for cell in ws['A']:
            if not (cell.value and isinstance(cell.value, str)):
                continue
            channel_id = cell.value.strip()
            hex_color = color_map.get(channel_id)
            if not hex_color:
                continue
            argb = _hex_to_argb(hex_color)
            cell.fill = PatternFill(patternType='solid', fgColor=argb)

    wb.save(dst_path)
