"""
元Excelをコピーし、UIで設定した区分色を座席セルと凡例セルに適用する。

- 割当済み座席セル → 区分色を適用
- 未割当座席セル   → 元の色を保持（変更しない）
- 凡例セル (A列)  → 区分色を適用
- セルテキスト    → 元の値を保持
"""
import shutil
import openpyxl
from openpyxl.styles import PatternFill, Font

from excel_importer import get_seat_cell_map

_CATEGORY_SHEET_NAME = '区分'


def _hex_to_argb(hex_color: str) -> str:
    """'#RRGGBB' → 'FFRRGGBB' (ARGB)"""
    h = hex_color.lstrip('#')
    if len(h) == 3:
        h = ''.join(c * 2 for c in h)
    return ('FF' + h).upper()


def _make_fill(hex_color: str) -> PatternFill:
    return PatternFill(patternType='solid', fgColor=_hex_to_argb(hex_color))


def _add_category_sheet(wb: openpyxl.Workbook, categories: list) -> None:
    """「区分」シートをワークブックに追加（既存なら上書き）。"""
    if _CATEGORY_SHEET_NAME in wb.sheetnames:
        del wb[_CATEGORY_SHEET_NAME]
    ws = wb.create_sheet(title=_CATEGORY_SHEET_NAME)

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 12

    bold = Font(bold=True)
    for hdr, col in (('区分ID', 1), ('区分名', 2), ('色コード', 3)):
        c = ws.cell(row=1, column=col, value=hdr)
        c.font = bold

    for r, cat in enumerate(categories, start=2):
        cat_id    = cat.get('id', '')
        cat_name  = cat.get('name', cat_id)
        hex_color = cat.get('color')

        ws.cell(row=r, column=1, value=cat_id)
        ws.cell(row=r, column=2, value=cat_name)
        ws.cell(row=r, column=3, value=hex_color or '')

        if hex_color:
            fill = _make_fill(hex_color)
            for col in (1, 2):
                ws.cell(row=r, column=col).fill = fill


def apply_seat_and_legend_colors(src_path: str, dst_path: str, data: dict) -> None:
    """
    src_path をコピーして dst_path に保存し、以下を適用する。
    - data['categories'] の色 → A列の凡例セル
    - data['sheets'][*]['seats'] の categoryId → 対応する座席セル

    data: exportFromServer が送る {version, categories, sheets} JSON
    """
    # カテゴリID → 色 のマップ
    color_map: dict[str, str] = {
        cat['id']: cat['color']
        for cat in data.get('categories', [])
        if cat.get('color')
    }

    # seat_id → (sheet_name, row, col)
    cell_map = get_seat_cell_map(src_path)

    shutil.copy2(src_path, dst_path)
    wb = openpyxl.load_workbook(dst_path)

    # 座席セルに色を適用（割当済みのみ）
    for sheet_name, sheet_data in data.get('sheets', {}).items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for seat_id, seat_info in sheet_data.get('seats', {}).items():
            cat_id = seat_info.get('categoryId')
            if not cat_id:
                continue  # 未割当: 元の色を保持
            hex_color = color_map.get(cat_id)
            if not hex_color:
                continue
            pos = cell_map.get(seat_id)
            if not pos:
                continue
            _, row, col = pos
            ws.cell(row=row, column=col).fill = _make_fill(hex_color)

    # 凡例セル (A列) に色を適用
    for ws in wb.worksheets:
        for cell in ws['A']:
            if not (cell.value and isinstance(cell.value, str)):
                continue
            hex_color = color_map.get(cell.value.strip())
            if hex_color:
                cell.fill = _make_fill(hex_color)

    _add_category_sheet(wb, data.get('categories', []))
    wb.save(dst_path)
