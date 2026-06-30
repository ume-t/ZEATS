"""
ZEATSのJSONデータを色付きExcelファイルに変換するモジュール。

入力: parse_excel()の戻り値（categories + sheets[].seats + sheets[].blocks）
出力: 座席ブロックごとに区分色が付いたExcelファイル

レイアウト（シートごとに縦積み）:
  [ブロック名]
  [座席][座席][座席]...  ← 列ごとに1行
  [座席][座席][座席]...
  （空白行）
  [次のブロック名]
  ...
"""

import re
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side


_THIN = Side(style='thin', color='CCCCCC')
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HEADER_FONT = Font(bold=True)
_CELL_ALIGN = Alignment(horizontal='center', vertical='center')


def _argb(hex_color: str) -> str:
    """'#RRGGBB' → 'FFRRGGBB'（openpyxlのARGB形式）"""
    return 'FF' + hex_color.lstrip('#').upper()


def _seat_number(seat_id: str) -> str:
    """seat_id の末尾 '{N}番' から座席番号を取り出す。"""
    m = re.search(r'_(\d+)番$', seat_id)
    return m.group(1) if m else seat_id


def _seat_label(seat_id: str, seat_info: dict) -> str:
    """セルに表示するテキスト。チケット番号 > 座席番号の優先順。"""
    ticket = seat_info.get('ticketNo')
    return ticket if ticket else _seat_number(seat_id)


def export_excel(data: dict, output_path: str) -> None:
    """
    JSONデータを色付きExcelファイルに出力する。

    Args:
        data: parse_excel()の戻り値。
              categories / sheets[].seats / sheets[].blocks を使用。
        output_path: 出力先ファイルパス（.xlsx）
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # カテゴリID → PatternFill のマップ
    cat_fills: dict[str, PatternFill] = {}
    for cat in data.get('categories', []):
        cat_fills[cat['id']] = PatternFill(
            fill_type='solid', fgColor=_argb(cat['color'])
        )

    for sheet_name, sheet_data in data.get('sheets', {}).items():
        ws = wb.create_sheet(title=sheet_name[:31])
        ws.column_dimensions['A'].width = 14
        seats = sheet_data['seats']
        blocks = sheet_data['blocks']

        current_row = 1
        for block in blocks:
            # ブロック名ヘッダー行
            header = ws.cell(row=current_row, column=1, value=block['name'])
            header.font = _HEADER_FONT
            current_row += 1

            # 座席行
            for row_seats in block['rows']:
                for col_i, seat_id in enumerate(row_seats, start=1):
                    seat_info = seats.get(seat_id, {})
                    cell = ws.cell(
                        row=current_row,
                        column=col_i,
                        value=_seat_label(seat_id, seat_info),
                    )
                    cell.alignment = _CELL_ALIGN
                    cell.border = _BORDER

                    cat_id = seat_info.get('categoryId')
                    if cat_id and cat_id in cat_fills:
                        cell.fill = cat_fills[cat_id]

                current_row += 1

            current_row += 1  # ブロック間の空白行

    wb.save(output_path)
    print(f"Excel出力: {output_path}")


if __name__ == '__main__':
    import sys
    import json
    from excel_importer import parse_excel

    if len(sys.argv) < 2:
        print("使い方: python excel_exporter.py <input.xlsx> [output.xlsx]")
        sys.exit(1)

    input_path = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else 'output_colored.xlsx'

    print(f"インポート中: {input_path}")
    data = parse_excel(input_path)

    total = sum(len(s['seats']) for s in data['sheets'].values())
    print(f"シート数: {len(data['sheets'])}, 合計座席数: {total}")

    export_excel(data, output_path)
